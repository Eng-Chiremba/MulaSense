from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
import csv
import json
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from accounting.models import Transaction, Category
from budget.models import BudgetCategory, Goal, GoalContribution
from .serializers import ReportRequestSerializer

# Dashboard Data Endpoints
@api_view(['GET'])
def dashboard_overview(request):
    user = request.user
    current_month = timezone.now().replace(day=1)
    
    # Financial Summary
    income = Transaction.objects.filter(
        user=user, transaction_type='income',
        transaction_date__gte=current_month, status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    expenses = Transaction.objects.filter(
        user=user, transaction_type='expense',
        transaction_date__gte=current_month, status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Budget Performance
    budget_categories = BudgetCategory.objects.filter(user=user, is_active=True)
    over_budget_count = sum(1 for cat in budget_categories if cat.percentage_used > 100)
    
    # Goals Progress
    active_goals = Goal.objects.filter(user=user, status='active')
    goals_on_track = sum(1 for goal in active_goals if goal.days_remaining > 0)
    
    # Recent Activity
    recent_transactions = Transaction.objects.filter(
        user=user, status='completed'
    ).order_by('-transaction_date')[:5]
    
    return Response({
        'financial_summary': {
            'monthly_income': income,
            'monthly_expenses': expenses,
            'balance': income - expenses,
            'savings_rate': round((income - expenses) / income * 100, 1) if income > 0 else 0
        },
        'budget_performance': {
            'total_categories': budget_categories.count(),
            'over_budget': over_budget_count,
            'categories': [{
                'name': cat.name,
                'percentage_used': cat.percentage_used,
                'status': 'over' if cat.percentage_used > 100 else 'on_track'
            } for cat in budget_categories[:5]]
        },
        'goals_summary': {
            'total_goals': active_goals.count(),
            'on_track': goals_on_track,
            'goals': [{
                'name': goal.name,
                'progress': goal.progress_percentage,
                'days_remaining': goal.days_remaining
            } for goal in active_goals[:3]]
        },
        'recent_activity': [{
            'description': t.description,
            'amount': t.amount,
            'type': t.transaction_type,
            'date': t.transaction_date,
            'category': t.category.name
        } for t in recent_transactions]
    })

@api_view(['GET'])
def financial_metrics(request):
    user = request.user
    period = request.GET.get('period', 'month')
    
    # Calculate date range
    now = timezone.now()
    if period == 'week':
        start_date = now - timedelta(days=7)
    elif period == 'year':
        start_date = now.replace(month=1, day=1)
    else:
        start_date = now.replace(day=1)
    
    transactions = Transaction.objects.filter(
        user=user, transaction_date__gte=start_date, status='completed'
    )
    
    # Category breakdown
    expense_categories = transactions.filter(transaction_type='expense').values(
        'category__name', 'category__color'
    ).annotate(total=Sum('amount')).order_by('-total')[:10]
    
    income_categories = transactions.filter(transaction_type='income').values(
        'category__name', 'category__color'
    ).annotate(total=Sum('amount')).order_by('-total')[:5]
    
    # Trends (compare with previous period)
    prev_start = start_date - (now - start_date)
    prev_transactions = Transaction.objects.filter(
        user=user, transaction_date__gte=prev_start,
        transaction_date__lt=start_date, status='completed'
    )
    
    current_expenses = transactions.filter(transaction_type='expense').aggregate(Sum('amount'))['amount__sum'] or 0
    prev_expenses = prev_transactions.filter(transaction_type='expense').aggregate(Sum('amount'))['amount__sum'] or 0
    expense_trend = ((current_expenses - prev_expenses) / prev_expenses * 100) if prev_expenses > 0 else 0
    
    return Response({
        'period': period,
        'expense_categories': list(expense_categories),
        'income_categories': list(income_categories),
        'trends': {
            'expense_change': round(expense_trend, 1),
            'direction': 'up' if expense_trend > 0 else 'down'
        }
    })

# Export Functionality
@api_view(['GET'])
def export_transactions_csv(request):
    user = request.user
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    transactions = Transaction.objects.filter(user=user, status='completed')
    if start_date and end_date:
        transactions = transactions.filter(
            transaction_date__date__gte=start_date,
            transaction_date__date__lte=end_date
        )
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="transactions.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Date', 'Description', 'Category', 'Type', 'Amount', 'Status'])
    
    for transaction in transactions.order_by('-transaction_date'):
        writer.writerow([
            transaction.transaction_date.strftime('%Y-%m-%d'),
            transaction.description,
            transaction.category.name,
            transaction.transaction_type,
            str(transaction.amount),
            transaction.status
        ])
    
    return response

@api_view(['GET'])
def export_budget_csv(request):
    user = request.user
    categories = BudgetCategory.objects.filter(user=user, is_active=True)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="budget.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Category', 'Budgeted Amount', 'Spent Amount', 'Remaining', 'Percentage Used', 'Period'])
    
    for category in categories:
        writer.writerow([
            category.name,
            str(category.budgeted_amount),
            str(category.spent_amount),
            str(category.remaining_amount),
            f"{category.percentage_used:.1f}%",
            category.period
        ])
    
    return response

@api_view(['GET'])
def export_financial_report_pdf(request):
    user = request.user
    start_date = request.GET.get('start_date', timezone.now().replace(day=1).date())
    end_date = request.GET.get('end_date', timezone.now().date())
    
    # Create PDF
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, height - 50, f"Financial Report - {start_date} to {end_date}")
    
    # Summary data
    transactions = Transaction.objects.filter(
        user=user, transaction_date__date__gte=start_date,
        transaction_date__date__lte=end_date, status='completed'
    )
    
    income = transactions.filter(transaction_type='income').aggregate(Sum('amount'))['amount__sum'] or 0
    expenses = transactions.filter(transaction_type='expense').aggregate(Sum('amount'))['amount__sum'] or 0
    
    y_position = height - 100
    p.setFont("Helvetica", 12)
    p.drawString(50, y_position, f"Total Income: ${income:,.2f}")
    p.drawString(50, y_position - 20, f"Total Expenses: ${expenses:,.2f}")
    p.drawString(50, y_position - 40, f"Net Balance: ${income - expenses:,.2f}")
    
    # Category breakdown
    y_position -= 80
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y_position, "Expense Categories:")
    
    categories = transactions.filter(transaction_type='expense').values(
        'category__name'
    ).annotate(total=Sum('amount')).order_by('-total')[:10]
    
    y_position -= 20
    p.setFont("Helvetica", 10)
    for category in categories:
        p.drawString(70, y_position, f"{category['category__name']}: ${category['total']:,.2f}")
        y_position -= 15
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="financial_report.pdf"'
    
    return response

@api_view(['GET'])
def export_balance_sheet_excel(request):
    user = request.user
    
    # Calculate balance sheet data
    current_date = timezone.now().date()
    
    # Assets (positive balances)
    cash_income = Transaction.objects.filter(
        user=user, transaction_type='income', status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    cash_expenses = Transaction.objects.filter(
        user=user, transaction_type='expense', status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    cash_balance = cash_income - cash_expenses
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    
    # Header styling
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=12)
    currency_format = '#,##0.00'
    
    # Title
    ws['A1'] = f"Balance Sheet as of {current_date}"
    ws['A1'].font = header_font
    ws.merge_cells('A1:C1')
    
    row = 3
    
    # ASSETS
    ws[f'A{row}'] = "ASSETS"
    ws[f'A{row}'].font = subheader_font
    row += 1
    
    ws[f'A{row}'] = "Current Assets:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    ws[f'B{row}'] = "Cash and Cash Equivalents"
    ws[f'C{row}'] = cash_balance
    ws[f'C{row}'].number_format = currency_format
    row += 1
    
    ws[f'B{row}'] = "Accounts Receivable"
    ws[f'C{row}'] = 0
    ws[f'C{row}'].number_format = currency_format
    row += 1
    
    total_current_assets = cash_balance
    ws[f'A{row}'] = "Total Current Assets"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = total_current_assets
    ws[f'C{row}'].number_format = currency_format
    ws[f'C{row}'].font = Font(bold=True)
    row += 2
    
    # LIABILITIES
    ws[f'A{row}'] = "LIABILITIES"
    ws[f'A{row}'].font = subheader_font
    row += 1
    
    ws[f'B{row}'] = "Accounts Payable"
    ws[f'C{row}'] = 0
    ws[f'C{row}'].number_format = currency_format
    row += 1
    
    total_liabilities = 0
    ws[f'A{row}'] = "TOTAL LIABILITIES"
    ws[f'A{row}'].font = subheader_font
    ws[f'C{row}'] = total_liabilities
    ws[f'C{row}'].number_format = currency_format
    ws[f'C{row}'].font = subheader_font
    row += 2
    
    # EQUITY
    ws[f'A{row}'] = "OWNER'S EQUITY"
    ws[f'A{row}'].font = subheader_font
    row += 1
    
    retained_earnings = cash_balance
    ws[f'B{row}'] = "Retained Earnings"
    ws[f'C{row}'] = retained_earnings
    ws[f'C{row}'].number_format = currency_format
    row += 1
    
    ws[f'A{row}'] = "TOTAL OWNER'S EQUITY"
    ws[f'A{row}'].font = subheader_font
    ws[f'C{row}'] = retained_earnings
    ws[f'C{row}'].number_format = currency_format
    ws[f'C{row}'].font = subheader_font
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="balance_sheet_{current_date}.xlsx"'
    
    return response
